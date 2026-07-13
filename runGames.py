import subprocess
import re
import time
from datetime import datetime
from collections import defaultdict

# Excel styling imports
import openpyxl
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Font, PatternFill, Alignment

def run_command(cmd):
    """Run command and return output"""
    print(f"[*] {cmd}")
    try:
        result = subprocess.run(cmd, shell=True, capture_output=True, text=True, timeout=30)
        return result.stdout
    except subprocess.TimeoutExpired:
        return ""

def extract_match_id(output):
    """Extract match ID from output"""
    match = re.search(r'[a-f0-9]{8}-[a-f0-9]{4}-[a-f0-9]{4}-[a-f0-9]{4}-[a-f0-9]{12}', output)
    if match:
        return match.group(0)
    return None

def extract_match_data(info_output):
    """
    Parses the completed match info specifically for the fcode CLI table layout.
    """
    team_a_match = re.search(r'Team A:\s+([^\(]+)', info_output)
    team_b_match = re.search(r'Team B:\s+([^\(]+)', info_output)
    
    team_a = team_a_match.group(1).strip() if team_a_match else "UnknownA"
    team_b = team_b_match.group(1).strip() if team_b_match else "UnknownB"
    
    team_a = team_a.split("<-")[0].strip()
    team_b = team_b.split("<-")[0].strip()
    
    my_team = "Oogway"
    opponent = team_b if my_team in team_a else team_a
        
    maps = {}
    
    table_lines = re.findall(r'│\s*\d+\s*│\s*([a-zA-Z0-9_]+)\s*│\s*([A-Za-z0-9_\s\(\)]+?)│', info_output)
    
    for map_name, winner_cell in table_lines:
        map_name = map_name.strip()
        winner_cell = winner_cell.strip()
        
        if my_team in winner_cell:
            maps[map_name] = "Win"
        elif "A" in winner_cell or "B" in winner_cell:
            maps[map_name] = "Loss"
            
    return opponent, maps

def main():
    team_ids = [
        "cb7fc856-8648-4b62-9b60-ddfe3311b31d",
        "a0dfa333-e015-4b7f-bd32-15ca56f6a1d8"
    ]

    # { opponent_team: { map_name: { 'wins': X, 'total': Y } } }
    stats_tracker = defaultdict(lambda: defaultdict(lambda: {'wins': 0, 'total': 0}))
    all_maps = set()       
    batch_number = 0
    total_matches_run = 0

    print("="*70)
    print("STARTING CONTINUOUS TRACKER WITH EXCEL HEATMAP")
    print("="*70)

    while True:
        batch_number += 1
        current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        print(f"\n\n{'='*70}")
        print(f"STARTING BATCH {batch_number} | TIME: {current_time}")
        print(f"{'='*70}")

        for i, team_id in enumerate(team_ids, 1):
            print(f"\n--- Match {i} of Batch {batch_number} ---")

            # Queue match
            cmd = f"fcode match unrated {team_id}"
            output = run_command(cmd)
            
            match_id = extract_match_id(output)
            if not match_id:
                print(f"[-] Could not find match ID in output. Skipping...")
                continue

            print(f"[+] Match ID: {match_id}")
            print("[*] Polling for match to complete...")

            info_output = ""
            
            # Poll loop
            while True:
                time.sleep(10)
                cmd_info = f"fcode match info {match_id}"
                info_output = run_command(cmd_info)
                
                lower_info = info_output.lower()
                if "status:  complete" in lower_info or "status: complete" in lower_info:
                    print(f"[+] Match {match_id} completed!")
                    break
                elif "status:  failed" in lower_info or "status: failed" in lower_info:
                    print(f"[-] Match {match_id} failed!")
                    break
                    
                print("    ...still queued/running.")

            # Extract Data
            opponent, map_scores = extract_match_data(info_output)
            
            if map_scores:
                total_matches_run += 1
                for map_name, result in map_scores.items():
                    all_maps.add(map_name)
                    stats_tracker[opponent][map_name]['total'] += 1
                    if result == "Win":
                        stats_tracker[opponent][map_name]['wins'] += 1

        # GENERATE STYLIZED EXCEL WITH HEATMAP
        print("\n" + "-"*70)
        print("UPDATING gameResults.xlsx WITH COLOR GRADIENT")
        print("-"*70)

        sorted_maps = sorted(list(all_maps))
        map_totals = {m: {'wins': 0, 'total': 0} for m in sorted_maps}
        sorted_opponents = sorted(stats_tracker.keys())

        # Initialize workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Win Rates"
        ws.views.sheetView[0].showGridLines = True

        # Header formatting
        header = ['team'] + sorted_maps + ['Total']
        ws.append(header)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        # Write data rows
        for opponent in sorted_opponents:
            row = [opponent]
            team_wins = 0
            team_total = 0
            
            for map_name in sorted_maps:
                map_stats = stats_tracker[opponent].get(map_name)
                if map_stats and map_stats['total'] > 0:
                    team_wins += map_stats['wins']
                    team_total += map_stats['total']
                    map_totals[map_name]['wins'] += map_stats['wins']
                    map_totals[map_name]['total'] += map_stats['total']
                    
                    row.append(round(map_stats['wins'] / map_stats['total'], 2))
                else:
                    row.append('N/A')
            
            # Weighted average total row margin calculation
            if team_total > 0:
                row.append(round(team_wins / team_total, 2))
            else:
                row.append('N/A')
            ws.append(row)

        # Bottom Row (Weighted map totals)
        bottom_row = ['Total']
        grand_wins = 0
        grand_total = 0
        for map_name in sorted_maps:
            m_stats = map_totals[map_name]
            if m_stats['total'] > 0:
                grand_wins += m_stats['wins']
                grand_total += m_stats['total']
                bottom_row.append(round(m_stats['wins'] / m_stats['total'], 2))
            else:
                bottom_row.append('N/A')
        
        # Grand Total cell calculation
        if grand_total > 0:
            bottom_row.append(round(grand_wins / grand_total, 2))
        else:
            bottom_row.append('N/A')
        ws.append(bottom_row)

        # Style the Total row specifically
        last_row_idx = ws.max_row
        for cell in ws[last_row_idx]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")

        # Dynamically size columns nicely
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        # Apply Color Scale Gradient Rule (0 = Dark Red, 0.5 = White/Yellow, 1 = Light Blue)
        # Ranges cover B2 down to the final total values
        max_col_letter = openpyxl.utils.get_column_letter(ws.max_column)
        color_range = f"B2:{max_col_letter}{last_row_idx}"
        
        color_scale = ColorScaleRule(
            start_type='num', start_value=0, start_color='FF8080', # Darker Red tint
            mid_type='num', mid_value=0.5, mid_color='FFFFE0',    # Soft Light Yellow
            end_type='num', end_value=1, end_color='ADD8E6'        # Light Blue
        )
        ws.conditional_formatting.add(color_range, color_scale)

        # Save workbook
        wb.save('gameResults.xlsx')
                
        print(f"[+] Updated gameResults.xlsx! Matches aggregated: {total_matches_run}")
        print("[*] Sleeping for 5 minutes before starting the next batch. Press Ctrl+C to stop.")
        try:
            time.sleep(300)
        except KeyboardInterrupt:
            print("\n[!] Script stopped by user. Final win rates saved in gameResults.xlsx.")
            break

if __name__ == "__main__":
    main()
